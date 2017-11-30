import openpyxl

""" Ojo que openpyxl hay que instalarlo, se puede con pip:

    pip install openpyxl
"""


class Documentador:


    def __init__(self, inmobiliaria):
        self.documento = openpyxl.Workbook()
        self.hoja1 = self.documento.create_sheet("Casas Vendidas")
        self.hoja2 = self.documento.create_sheet("Casas Sin Vender")
        self.hoja3 = self.documento.create_sheet("Precios en el tiempo")
        self.inmobiliaria = inmobiliaria
        self.columnas = iter(
            ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V',
             'W', 'X', 'Y', 'Z', 'AA', 'AB'])
        # Se definen los titulos de la hoja 1 y un contador
        self.hoja1['A1'] = 'Tiempo'
        self.hoja1['B1'] = 'Casa'
        self.hoja1['C1'] = 'Precio'
        self.hoja1['D1'] = 'Descuento'
        self.hoja1['E1'] = 'Cluster'
        self.contador = 2
        # Se definen los titulos de la hoja 3
        self.hoja3['A1'] = 'Casa'
        for i in range(100):
            self.hoja3['A' + str(i + 2)] = 'Casa ' + str(i + 1)

    def casa_vendida(self, casa, tiempo, tipo_cliente):
        self.hoja1['A' + str(self.contador)] = tiempo // (24)
        self.hoja1['B' + str(self.contador)] = casa.identificador
        self.hoja1['C' + str(self.contador)] = casa.precio
        self.hoja1['D' + str(self.contador)] = casa.precio_inicial - casa.precio
        self.hoja1['E' + str(self.contador)] = tipo_cliente
        self.contador += 1

    def cambio_precio(self, semana):
        col = next(self.columnas)
        self.hoja3[col + '1'] = "Semana " + str(semana)
        cont = 2
        for i in self.inmobiliaria.casas:
            self.hoja3[col + str(cont)] = i.precio
            cont += 1

    def fin_simulacion(self):
        i = 2
        for k in range(100):
            if not self.inmobiliaria.casas[k].vendida:
                casa = [str(self.hoja2['A' + str(i)].value), 1, 1, 0, 1, 0]
                self.hoja2['A' + str(i)] = self.inmobiliaria.casas[k].identificador
                columnas = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']
                for j in range(9):
                    self.hoja2[columnas[j] + str(i)] = self.inmobiliaria.casas[k].atributos[j + 5]
                i += 1

        self.documento.save('Resultados.xlsx')


class Docu2:

    '''es un simil al anterior Documentador'''
    def __init__(self, precios_promedio_casas_vendidas):
        self.precios_promedio_casas_vendidas = precios_promedio_casas_vendidas
        self.documento = openpyxl.Workbook()
        self.hoja1 = self.documento.create_sheet("Precio promedio casas vendidas")
        self.hoja1['A1'] = 'Casa'
        for i in range(100):
            self.hoja1['A' + str(i + 2)] = 'Casa ' + str(i + 1)
        self.hoja1['B1'] = 'Precio promedio'
        self.hoja1['C1'] = 'Varianza'
        for i in range(100):
            self.hoja1['B' + str(i + 2)] = self.precios_promedio_casas_vendidas[i][0]
            self.hoja1['C' + str(i + 2)] = self.precios_promedio_casas_vendidas[i][1]
        self.documento.save('Precios promedio casas vendidas.xlsx')

    def cantidad_personas_no_compra(self, lista_cantidad_personas_no_compra, varianza, simulaciones):
        self.docu2 = openpyxl.Workbook()
        self.hoja1 = self.docu2.create_sheet("Personas no compran")
        self.hoja1['A1'] = 'Simulacion'
        self.hoja1['B1'] = 'Cantidad personas no compra'
        self.hoja1['D1'] = 'Cantidad promedio personas no compra'
        self.hoja1['E1'] = 'Varianza'
        for i in range(simulaciones):
            self.hoja1['A' + str(i + 2)] = str(i + 1)
            self.hoja1['B' + str(i + 2)] = lista_cantidad_personas_no_compra[i]
        self.hoja1['D2'] = sum(lista_cantidad_personas_no_compra)/len(lista_cantidad_personas_no_compra)
        self.hoja1['E2'] = varianza
        self.docu2.save('Personas no compran.xlsx')

    def cantidad_personas_compra(self, lista_casas_vendidas, varianza, simulaciones):
        self.docu3 = openpyxl.Workbook()
        self.hoja0 = self.docu3.create_sheet("Personas compran")
        self.hoja0['A1'] = 'Simulacion'
        self.hoja0['B1'] = 'Cantidad personas compra'
        self.hoja0['D1'] = 'Cantidad promedio personas compra'
        self.hoja0['E1'] = 'Varianza'
        for i in range(simulaciones):
            self.hoja0['A' + str(i + 2)] = str(i + 1)
            self.hoja0['B' + str(i + 2)] = lista_casas_vendidas[i]
        self.hoja0['D2'] = sum(lista_casas_vendidas) / len(lista_casas_vendidas)
        self.hoja0['E2'] = varianza
        self.docu3.save('Personas compran.xlsx')

    def factor_precios_por_periodo(self, matriz_factor_precios_para_cada_casa_por_periodo):
        self.docu4 = openpyxl.Workbook()
        self.hoja0 = self.docu4.create_sheet("Factor de precio promedio")
        self.hoja1 = self.docu4.create_sheet("Varianza factor de precio")
        columnas = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T',
                         'U', 'V', 'W', 'X', 'Y', 'Z']
        self.hoja0['A1'] = 'Casa'
        self.hoja1['A1'] = 'Casa'
        for i in range(100):
            self.hoja0['A' + str(i + 2)] = str(i + 1)
            self.hoja1['A' + str(i + 2)] = str(i + 1)
        for col in columnas:
            if col == 'B':
                self.hoja0[col + str(1)] = 'Semana 0'
                self.hoja1[col + str(1)] = 'Semana 0'
            elif col == 'C':
                self.hoja0[col + str(1)] = 'Semana 7'
                self.hoja1[col + str(1)] = 'Semana 7'
            else:
                indice = columnas.index(col)
                self.hoja0[col + str(1)] = 'Semana ' + str(7 + (indice - 1) * 4)
                self.hoja1[col + str(1)] = 'Semana ' + str(7 + (indice - 1) * 4)
        for casa in range(100):
            for periodo in range(25):
                self.hoja0[columnas[periodo] + str(casa + 2)] = \
                    matriz_factor_precios_para_cada_casa_por_periodo[casa][periodo][0]
                self.hoja1[columnas[periodo] + str(casa + 2)] = \
                    matriz_factor_precios_para_cada_casa_por_periodo[casa][periodo][1]
        self.docu4.save('Factor precio.xlsx')

    def precios_promedio_por_periodo(self, precio_promedio_y_varianza_por_casa_y_periodo):
        self.docu5 = openpyxl.Workbook()
        self.hoja0 = self.docu5.create_sheet("Precio promedio por periodo")
        self.hoja1 = self.docu5.create_sheet("Varianza de precio promedio")
        columnas = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T',
                    'U', 'V', 'W', 'X', 'Y', 'Z']
        self.hoja0['A1'] = 'Casa'
        self.hoja1['A1'] = 'Casa'
        for i in range(100):
            self.hoja0['A' + str(i + 2)] = str(i + 1)
            self.hoja1['A' + str(i + 2)] = str(i + 1)
        for col in columnas:
            if col == 'B':
                self.hoja0[col + str(1)] = 'Semana 0'
                self.hoja1[col + str(1)] = 'Semana 0'
            elif col == 'C':
                self.hoja0[col + str(1)] = 'Semana 7'
                self.hoja1[col + str(1)] = 'Semana 7'
            else:
                indice = columnas.index(col)
                self.hoja0[col + str(1)] = 'Semana ' + str(7 + (indice - 1) * 4)
                self.hoja1[col + str(1)] = 'Semana ' + str(7 + (indice - 1) * 4)
        for casa in range(100):
            for periodo in range(25):
                self.hoja0[columnas[periodo] + str(casa + 2)] = \
                    precio_promedio_y_varianza_por_casa_y_periodo[casa][periodo][0]
                self.hoja1[columnas[periodo] + str(casa + 2)] = \
                    precio_promedio_y_varianza_por_casa_y_periodo[casa][periodo][1]
        self.docu5.save('Precios promedio por periodo.xlsx')