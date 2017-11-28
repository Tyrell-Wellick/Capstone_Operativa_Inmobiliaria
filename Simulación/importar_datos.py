import openpyxl
import numpy
from itertools import groupby
from operator import itemgetter
""" Ojo que openpyxl hay que instalarlo, se puede con pip:
    
    pip install openpyxl
"""

def importar_casas(alfa):
    doc = openpyxl.load_workbook('Datos.xlsx')
    hoja = doc.get_sheet_by_name("DatosCasas")
    casas = [] #Lista que tendra los datos de cada casa
    for i in range(2, 102):
        casa = [str(hoja['A' + str(i)].value), 1, 1, 0, 1, 0] 
        """Los primeros atributos de cada
        casa son los atributos del condominio"""
        columnas = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']
        for j in columnas:
            casa.append(int(hoja[j + str(i)].value))

            """ Si llegamos a poner los precios iniciales de las casas
            en el excel habría que descomentar la siguiente linea para
            que los importe. Lo mismo en la definicion de la inmobiliaria"""
            #casa.append(int(hoja['K' + str(i)].value))

        casas.append(casa)
    casas_in=casas
    
    importantes = [[i[0],(1, *i[6:10], i[11], i[13])] for i in casas_in]
    
    
    
    seq = importantes
    seq.sort(key = itemgetter(1))
    groups = groupby(seq, itemgetter(1))
    grupos_casas = [[item[0] for item in data] for (key, data) in groups]
    
    
    for grupo in grupos_casas:
        simb = grupo[0]
        for casa in importantes:
            if casa[0] == simb:
                grupo.append(casa[1:])

    doc2 = openpyxl.load_workbook('Datos.xlsx')
    hoja2 = doc2.get_sheet_by_name("Información Histórica")
    casas=[]
    counter = -1
    for k in hoja2.iter_rows():
        if counter <= 0:
            pass
        else:
            valores= (k[1], k[2], k[3], k[8], k[9], k[10], k[11],
                      k[13], k[15], k[17], k[18])
            
            casas.append([i.value for i in valores])
        counter += 1
    
    
    casas2 = []
    for index, i in  enumerate(casas):
        casas2.append([i[:2],i[2:9],  float(i[9]), i[10]])
        
    
    
    seq = casas2
    seq.sort(key = itemgetter(1))
    groups = groupby(seq, itemgetter(1))
    casas_hist = [[item for item in data] for (key, data) in groups]
    
    
    
    contador = 0
    lista_precios = []
    for grupo in grupos_casas:
        for grupo2 in casas_hist:
            if list(grupo[-1][0]) == grupo2[0][1]:
                contador += 1
                inst = []
                for instance in grupo2:
                    inst.append((instance[2],instance[3]))
                precios = [i[0]+alfa*((i[0])*(100-i[1])/100) for i in inst]            
    
                break
        lista_precios.append([grupo[:-1],
                              numpy.mean(precios)])
      
    lista_casas=[]
    for precio in lista_precios:
        for casa in casas_in:
            if casa[0] in precio[:-1][0]:
                lista_casas.append([int(casa[0].split(' ')[1]), precio[-1]]) 
                pass
    
    lista_casas.sort(key = itemgetter(0))
    
    for index, i in enumerate(lista_casas):
        casas_in[index].append(i[1])

    return casas_in

            
            